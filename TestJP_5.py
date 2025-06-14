import datetime
import re
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

import openpyxl

# Excelファイルのパスとシート名を設定
FILE_PATH = "JP_Stocks.xlsx"
SHEET_NAME = "Price"


def load_sheet(path: str, sheet: str):
    """Excel ファイルとシートを読み込む"""
    wb = openpyxl.load_workbook(path)
    return wb, wb[sheet]


def parse_dates(ws):
    """1 行目の yy/mm 形式から月末日を取得"""
    dates = []
    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if not value:
            break
        m = re.match(r"(\d{2})/(\d{2})", str(value))
        if m:
            yy, mm = map(int, m.groups())
            year = 2000 + yy
            # 月末日を計算し、休日なら直前の平日にずらす
            d = datetime.date(year, mm, 1).replace(day=28) + datetime.timedelta(days=4)
            last_day = d - datetime.timedelta(days=d.day)
            while last_day.weekday() >= 5:  # 5=土,6=日
                last_day -= datetime.timedelta(days=1)
            dates.append((col, last_day))
    return dates


def parse_tickers(ws):
    """1 列目から銘柄コードを取得"""
    tickers = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        if code:
            tickers.append((row, str(code).strip()))
    return tickers


def init_driver():
    """Chrome WebDriver を初期化"""
    options = Options()
    # options.add_argument('--headless')  # 必要に応じて有効化
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(options=options)
    return driver


def fetch_price(driver, code: str, date: datetime.date):
    """指定日付の基準価額を取得"""
    url = f"https://finance.yahoo.co.jp/quote/{code}/history"
    driver.get(url)

    try:
        # プルダウンから "月間" を選択
        select_elem = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[2]/div/div[4]/div[2]/div/div[2]/div[1]/div/div/div/select"))
        )
        Select(select_elem).select_by_value("monthly")
        time.sleep(1)

        # 日付を入力（readonly 属性を外して値を直接設定）
        start_input = driver.find_element(By.ID, "historyTermFromDate")
        end_input = driver.find_element(By.ID, "historyTermToDate")

        date_str = date.strftime("%Y/%m/%d")
        for elem in (start_input, end_input):
            driver.execute_script("arguments[0].removeAttribute('readonly');", elem)
            driver.execute_script("arguments[0].value = arguments[1];", elem, date_str)

        # 表示ボタンを押下
        display_btn = driver.find_element(By.XPATH, "//a[@class='button__1hp- primary__3K8A  radius__1Hzm isBold__1O2x' and text()='表示']")
        display_btn.click()

        # テーブル表示を待機
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
        )
        time.sleep(1)

        # 終値セルから数値取得
        price_cell = driver.find_element(By.CSS_SELECTOR, "table tbody tr td:nth-child(2)")
        value_str = price_cell.text.strip().replace(",", "")
        return int(float(value_str))
    except Exception as e:
        print(f"{code} {date}: データ取得失敗 - {e}")
        return None


def main():
    wb, ws = load_sheet(FILE_PATH, SHEET_NAME)
    dates = parse_dates(ws)
    tickers = parse_tickers(ws)
    driver = init_driver()

    for row, code in tickers:
        for col, dt in dates:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                continue  # 既に値がある場合はスキップ
            price = fetch_price(driver, code, dt)
            if price is not None:
                cell.value = price
                print(f"{code} {dt}: {price}円")
            time.sleep(1)  # アクセス過多防止

    driver.quit()
    wb.save(FILE_PATH)
    print("✅ 全データ取得完了＆Excel保存！")


if __name__ == "__main__":
    main()
